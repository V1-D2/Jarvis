import openpyxl as xl
import Weather
from Weather import weatherPl



typesOfClouthese = {"hot":["Tshort", "Short", "Boots"],
                    "warm":["Hudi", "Trausers", "Boots"],
                    "cold":["Jaket", "Hudi", "Trausers", "Boots"]}
namesOfSheets = {"Jaket" : 0, "Hudi" : 1, "Trausers" : 2, "Boots" : 3, "Shorts" : 4, "Tshorts" : 5, "Termo" : 6, "Vetrovka" : 7, "Coft" : 8}

neededClouthes = {}
sportClouses = {}

def whatToTakeOnDependingOnTemp(temp, neededClouthes):
    if(temp > -30 and temp< 13):
        ar = typesOfClouthese["cold"]
        for i in ar:
            neededClouthes[i] = []
    elif(temp>=13 and temp < 23 ):
        ar = typesOfClouthese["warm"]
        for i in ar:
            neededClouthes[i] = []
    else:
        ar = typesOfClouthese["hot"]
        for i in ar:
            neededClouthes[i] = []
    return neededClouthes


def goToStreet(weatherF):
    global neededClouthes
    temp = weatherF["temp"]['temp']
    humidity = weatherF["humidity"]
    rain = weatherF["willItBeRaine"]
    snow = weatherF["willItBeSnowe"]

    if(humidity > 90): rain = True
    if (rain):rain = 1
    else:rain = 0
    if (snow):snow = 1
    else:snow = 0

    neededClouthes = whatToTakeOnDependingOnTemp(int(temp), neededClouthes)

    listOfClouses = whichClothesToDress(int(temp), rain, snow, neededClouthes)

    neededClouthes = {}
    return listOfClouses


def whichClothesToDress(temp, rain, snow, neededClouthes):
    wb = xl.load_workbook('CurrentClouses.xlsx')

    for nameOfSheet in neededClouthes.keys():
        sheet = wb[f'{nameOfSheet}']
        numberOfSheet = namesOfSheets[nameOfSheet]
        wb.active = numberOfSheet

        maxRow = sheet.max_row

        for i in range(2, maxRow):
            maxTemp = int(sheet.cell(row=i, column=2).value)
            if(maxTemp>temp):
                minTemp = int(sheet.cell(row = i, column = 3).value)
                if(minTemp<temp):
                    forSnow = int(sheet.cell(row = i, column = 4).value)
                    if(forSnow>=snow):
                        forRaine = int(sheet.cell(row = i, column = 5).value)
                        if(forRaine>=rain):
                            isAvailable = int(sheet.cell(row = i, column = 7).value)
                            if(isAvailable == 1):
                                neededClouthes[nameOfSheet].append(sheet.cell(row = i, column = 1).value)

    return neededClouthes







#Sport

def whichClothesToDressForSport(temp, rain, sportClouses, openOrClosed):
    wb = xl.load_workbook('CurrentClouses.xlsx')

    for nameOfSheet in sportClouses.keys():
        sheet = wb[f'{nameOfSheet}']
        numberOfSheet = namesOfSheets[nameOfSheet]
        wb.active = numberOfSheet

        if(openOrClosed == "closed"):
            temp = 23
            rain = 0
        maxRow = sheet.max_row
        for i in range(2, maxRow + 1):
            maxTemp = sheet.cell(row=i, column=2).value
            if(maxTemp>=temp):
                minTemp = sheet.cell(row = i, column = 3).value
                if(minTemp<=temp):
                    forRaine = sheet.cell(row = i, column = 5).value
                    if(forRaine>=rain):
                        forSport = sheet.cell(row = i, column = 6).value
                        if(forSport == 1):
                            isAvailable = int(sheet.cell(row=i, column=7).value)
                            if (isAvailable == 1):
                                sportClouses[nameOfSheet].append(sheet.cell(row = i, column = 1).value)
    return sportClouses



def goToSport(weatherF, closedOrNot):
    global sportClouses
    temp = weatherF["temp"]['temp']
    humidity = weatherF["humidity"]
    rain = weatherF["willItBeRaine"]
    snow = weatherF["willItBeSnowe"]

    if (humidity > 90): rain = True
    if (rain): rain = 1
    else: rain = 0
    if (snow): snow = 1
    else: snow =0

    if(closedOrNot == "closed"):
        sportClouses = closedSport(sportClouses)
    else:
        sportClouses = openedSport(weatherF, sportClouses)

    listOfSportClouses = whichClothesToDressForSport(temp, rain, sportClouses ,closedOrNot)
    sportClouses = {}
    return listOfSportClouses



def closedSport(sportClouses):
    sportClouses["Shorts"] = []
    sportClouses["Tshorts"] = []
    sportClouses["Boots"] = []

    return sportClouses

def openedSport(forecast, sportClouses):
    sportClouses["Shorts"] = []
    sportClouses["Tshorts"] = []
    sportClouses["Boots"] = []

    if(forecast["temp"]['temp']<15) or forecast["temp"]['temp']<17 and forecast["wind"]['speed']>15:
        sportClouses["Coft"] = []
    if(forecast["willItBeRaine"] or forecast["humidity"]>90):
        sportClouses["Vetrovka"] = []
        sportClouses["Termo"] = []
    if(forecast["temp"]['temp']<15):
        sportClouses["Termo"] = []

    return sportClouses


def nowIsNotAvailable(clothe):
    wb = xl.load_workbook("CurrentClouses.xlsx")

    for sheetName in namesOfSheets.keys():
        sheet = wb[f'{sheetName}']
        numberOfSheet = namesOfSheets[sheetName]
        wb.active = numberOfSheet

        maxRow = sheet.max_row
        for i in range(2, maxRow + 1):
            nameOfClothe = sheet.cell(row = i, column = 1).value
            if(str(nameOfClothe).lower() == clothe):
                sheet.cell(row = i, column = 7).value = 0
                break

    wb.save(filename="CurrentClouses.xlsx")

def nowIsAvailable(clothe):
    wb = xl.load_workbook("CurrentClouses.xlsx")

    for sheetName in namesOfSheets.keys():
        sheet = wb[f'{sheetName}']
        numberOfSheet = namesOfSheets[sheetName]
        wb.active = numberOfSheet

        maxRow = sheet.max_row
        for i in range(2, maxRow + 1):
            nameOfClothe = sheet.cell(row=i, column=1).value
            if (nameOfClothe.lower() == clothe):
                sheet.cell(row=i, column=7).value = 1
                break
    wb.save(filename="CurrentClouses.xlsx")





